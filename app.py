from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash, jsonify, send_file
import os
import pandas as pd
from datetime import datetime
from back import process_files
import warnings
import zipfile
import io
TEMPLATE_STORAGE = "template_files"
TEMPLATE_PATH = os.path.join(TEMPLATE_STORAGE, "current_template.xlsx")
os.makedirs(TEMPLATE_STORAGE, exist_ok=True)
FILLED_TEMPLATE_PATH = os.path.join(TEMPLATE_STORAGE, "filled_template.xlsx")

warnings.filterwarnings("ignore", category=DeprecationWarning)

# LangChain/Transformers Setup (rest as before)
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from transformers import AutoModelForSeq2SeqLM, AutoTokenizer, pipeline
from langchain_community.llms import HuggingFacePipeline
from langchain.chains import RetrievalQA
from langchain.prompts import PromptTemplate
from langchain_community.document_loaders import PyPDFLoader
import torch

UPLOAD_FOLDER = "uploads"
EXCEL_FOLDER = "extracted_excels"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)

EMBEDDING_MODEL = "sentence-transformers/all-MiniLM-L6-v2"
FLAN_MODEL = "google/flan-t5-base"

PROMPT = PromptTemplate(
    template="""
You are an intelligent assistant that answers questions based only on the provided context (a small excerpt of an invoice PDF).
- Only output the exact value **for the field the user asks about** (for instance, if asked 'what is the order number?' answer just the order number; do not repeat the question, do not provide unrelated lines).
- If not found, reply with: [Not found].
Context:
{context}
Question: {question}
Answer (ONLY the exact value; no extra explanation):
""",
    input_variables=["context", "question"]
)

def answer_question_per_pdf(question, pdf_folder):
    results = []
    for fname in os.listdir(pdf_folder):
        if not fname.lower().endswith('.pdf'):
            continue
        print(f"Processing PDF: {fname}")
        fullpath = os.path.join(pdf_folder, fname)
        try:
            loader = PyPDFLoader(fullpath)
            docs = loader.load()
            splitter = CharacterTextSplitter(chunk_size=300, chunk_overlap=50, separator='\n')
            chunks = splitter.split_documents(docs)
            embeddings = HuggingFaceEmbeddings(model_name=EMBEDDING_MODEL)
            vectordb = FAISS.from_documents(chunks, embeddings)
            retriever = vectordb.as_retriever(search_type="similarity", search_kwargs={"k": 3})
            llm = load_flan_llm()
            qa_chain = RetrievalQA.from_chain_type(
                llm=llm,
                chain_type="stuff",
                retriever=retriever,
                chain_type_kwargs={"prompt": PROMPT},
                return_source_documents=True
            )
            answer = qa_chain({"query": question})["result"].strip()
            if not answer or answer.lower().startswith("i don't know"):
                results.append(f"{fname}: [Not found]")
            else:
                results.append(f"{fname}: {answer}")
        except Exception as e:
            results.append(f"{fname}: [Error: {str(e)}]")
    return results

def load_flan_llm():
    tokenizer = AutoTokenizer.from_pretrained(FLAN_MODEL)
    model = AutoModelForSeq2SeqLM.from_pretrained(FLAN_MODEL)
    pipe = pipeline(
        "text2text-generation",
        model=model,
        tokenizer=tokenizer,
        max_new_tokens=150,
        device=0 if torch.cuda.is_available() else -1,
    )
    return HuggingFacePipeline(pipeline=pipe)

def get_file_list():
    files = []
    for filename in os.listdir(UPLOAD_FOLDER):
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        size = os.path.getsize(filepath)
        timestamp = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%b %d, %Y')
        files.append({
            'name': filename,
            'size': f"{size / 1024:.2f} KB",
            'timestamp': timestamp
        })
    return sorted(files, key=lambda x: x['timestamp'], reverse=True)

def get_excel_list():
    files = []
    if os.path.exists(EXCEL_FOLDER):
        for filename in os.listdir(EXCEL_FOLDER):
            filepath = os.path.join(EXCEL_FOLDER, filename)
            size = os.path.getsize(filepath)
            timestamp = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%b %d, %Y')
            try:
                with pd.ExcelFile(filepath) as excel_file:
                    sheet_count = len(excel_file.sheet_names)
                    sheet_names = excel_file.sheet_names
                    sheets = f"{sheet_count} Sheet{'s' if sheet_count != 1 else ''}"
            except:
                sheets = "1 Sheet"
                sheet_names = ["Sheet1"]
            files.append({
                'name': filename,
                'size': f"{size / 1024:.2f} KB",
                'timestamp': timestamp,
                'sheets': sheets,
                'sheet_names': sheet_names
            })
    return sorted(files, key=lambda x: x['timestamp'], reverse=True)

app = Flask(__name__)
app.secret_key = 'supersecret'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXCEL_FOLDER'] = EXCEL_FOLDER

# ------------ MAIN ROUTES ---------------

@app.route('/')
def dashboard():
    uploaded_files = get_file_list()
    excel_files = get_excel_list()
    stats = {
        'total_uploads': len(uploaded_files),
        'total_extractions': len(excel_files),
        'success_rate': round((len(excel_files) / len(uploaded_files) * 100) if uploaded_files else 0, 1),
        'recent_files': uploaded_files[:5]
    }
    return render_template("dashboard.html", stats=stats,
                           uploaded_files=uploaded_files,
                           excel_files=excel_files)

@app.route('/extraction')
def extraction():
    uploaded_files = get_file_list()
    excel_files = get_excel_list()
    return render_template("extraction.html",
                           uploaded_files=uploaded_files,
                           excel_files=excel_files)

@app.route('/visualization')
def visualization():
    excels = get_excel_list()
    uploaded_files = get_file_list()
    return render_template("visualization.html",
                           excel_files=excels,
                           uploaded_files=uploaded_files)

@app.route('/directory')
def directory():
    uploaded_files = get_file_list()
    excel_files = get_excel_list()
    return render_template("directory.html",
                           uploaded_files=uploaded_files,
                           excel_files=excel_files)

@app.route('/output_excels')
def output_excels():
    excel_files = get_excel_list()
    return render_template('output_excels.html', excel_files=excel_files)

# @app.route('/templates')
# def templates_page():
#     current_template = None
#     if os.path.exists(TEMPLATE_PATH):
#         current_template = os.path.basename(TEMPLATE_PATH)
#     return render_template('templates.html', current_template=current_template)

from openpyxl import load_workbook
import pickle
from sklearn.metrics.pairwise import cosine_similarity
import pandas as pd
import os

# def fill_template_all_sheets(output_file, save_path):
#     # Load extracted data
#     df_output = pd.read_excel(output_file)
#     extracted_headers = list(df_output.columns)

#     # Load template
#     wb = load_workbook(TEMPLATE_PATH)
#     ws_names = wb.sheetnames

#     # Load model
#     with open("cosine_model.pkl", "rb") as f:
#         model_cosine = pickle.load(f)
#     with open("cosine_data.pkl", "rb") as f:
#         cosine_data = pickle.load(f)
#     extracted_embeds = model_cosine.encode(extracted_headers)

#     # Loop through sheets
#     for ws_name in ws_names:
#         ws = wb[ws_name]
#         template_headers = [cell.value for cell in ws[1] if cell.value is not None]
#         if not template_headers:
#             continue
#         template_embeds = model_cosine.encode(template_headers)
#         col_indexes = {header: idx+1 for idx, header in enumerate(template_headers)}
#         mapping = {}
#         for i, ext_emb in enumerate(extracted_embeds):
#             sims = cosine_similarity([ext_emb], template_embeds)[0]
#             idx_best = sims.argmax()
#             if sims[idx_best] > 0.7:
#                 mapping[extracted_headers[i]] = template_headers[idx_best]
#         for extracted_col, template_col in mapping.items():
#             if extracted_col in df_output.columns and template_col in col_indexes:
#                 col_idx = col_indexes[template_col]
#                 for row_idx, value in enumerate(df_output[extracted_col], start=2):
#                     ws.cell(row=row_idx, column=col_idx, value=value)

#     wb.save(save_path)
#     return save_path
def fill_template_all_sheets(output_file, save_path):
    df_output = pd.read_excel(output_file)
    extracted_headers = list(df_output.columns)
    wb = load_workbook(FILLED_TEMPLATE_PATH)  # use always the live file!
    ws_names = wb.sheetnames
    with open("cosine_model.pkl", "rb") as f:
        model_cosine = pickle.load(f)
    with open("cosine_data.pkl", "rb") as f:
        cosine_data = pickle.load(f)
    extracted_embeds = model_cosine.encode(extracted_headers)
    for ws_name in ws_names:
        ws = wb[ws_name]
        template_headers = [cell.value for cell in ws[1] if cell.value is not None]
        if not template_headers:
            continue
        template_embeds = model_cosine.encode(template_headers)
        col_indexes = {header: idx+1 for idx, header in enumerate(template_headers)}
        mapping = {}
        for i, ext_emb in enumerate(extracted_embeds):
            sims = cosine_similarity([ext_emb], template_embeds)[0]
            idx_best = sims.argmax()
            if sims[idx_best] > 0.7:
                mapping[extracted_headers[i]] = template_headers[idx_best]
        for extracted_col, template_col in mapping.items():
            if extracted_col in df_output.columns and template_col in col_indexes:
                col_idx = col_indexes[template_col]
                # Find the first empty cell in this column after the header
                row_idx = 2
                while ws.cell(row=row_idx, column=col_idx).value not in (None, '',):
                    row_idx += 1
                # Write each new value below the last
                for value in df_output[extracted_col]:
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    row_idx += 1
    wb.save(FILLED_TEMPLATE_PATH)
    return FILLED_TEMPLATE_PATH

# --------- DOWNLOAD ALL EXCELS (ONE ROUTE ONLY, CORRECT NAME) ---------

@app.route('/download_all_excels')
def download_all_excels():
    """Zips all Excel files in EXCEL_FOLDER and serves them for download."""
    memory_file = io.BytesIO()
    files_found = False
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for filename in os.listdir(EXCEL_FOLDER):
            file_path = os.path.join(EXCEL_FOLDER, filename)
            if os.path.isfile(file_path):
                zf.write(file_path, arcname=filename)
                files_found = True
    memory_file.seek(0)
    if not files_found:
        flash("No Excel files to download.", "warning")
        return redirect(request.referrer or url_for('directory'))
    return send_file(
        memory_file, 
        download_name="all_excels.zip", 
        as_attachment=True, 
        mimetype="application/zip"
    )

# --------- OUTPUT EXCELS SECTION ENDPOINTS ---------

@app.route('/excel/<filename>')  # Download endpoint
def excel_file(filename):
    return send_from_directory(EXCEL_FOLDER, filename, as_attachment=True)

@app.route('/preview_excel/<filename>')
def preview_excel(filename):
    filepath = os.path.join(EXCEL_FOLDER, filename)
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(filepath)
        sheet_names = workbook.sheetnames
        sheet_data = {}
        for sheet_name in sheet_names:
            try:
                worksheet = workbook[sheet_name]
                data = []
                headers = []
                first_row = True
                for row in worksheet.iter_rows():
                    row_data = []
                    for cell in row:
                        cell_value = cell.value if cell.value is not None else ""
                        row_data.append(str(cell_value))
                    if first_row:
                        headers = row_data
                        first_row = False
                    else:
                        if any(cell.strip() for cell in row_data if cell):
                            data.append(row_data)
                if data:
                    df = pd.DataFrame(data, columns=headers)
                else:
                    df = pd.DataFrame(columns=headers)
                sheet_data[sheet_name] = df
            except Exception as sheet_error:
                sheet_data[sheet_name] = pd.DataFrame({
                    'Error': [f'Could not read sheet {sheet_name}: {str(sheet_error)}']
                })
        workbook.close()
        return render_template("excel_preview.html",
                               filename=filename,
                               sheet_data=sheet_data,
                               sheet_names=sheet_names)
    except Exception as e:
        return render_template("excel_preview.html",
                               filename=filename,
                               error=f"Could not load Excel file: {str(e)}")

@app.route('/api/excel_sheet/<filename>/<sheet_name>')
def get_excel_sheet_data(filename, sheet_name):
    try:
        filepath = os.path.join(EXCEL_FOLDER, filename)
        import openpyxl
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            data = []
            headers = []
            first_row = True
            for row in worksheet.iter_rows():
                row_data = []
                for cell in row:
                    cell_value = cell.value if cell.value is not None else ""
                    row_data.append(str(cell_value))
                if first_row:
                    headers = row_data
                    first_row = False
                else:
                    if any(cell.strip() for cell in row_data if cell):
                        data.append(row_data)
            if data:
                df = pd.DataFrame(data, columns=headers)
            else:
                df = pd.DataFrame(columns=headers)
            workbook.close()
            html_table = df.to_html(classes='excel-table', table_id='sheet-table', escape=False, index=False)
            return jsonify({
                'success': True,
                'html': html_table,
                'records': len(df),
                'columns': len(df.columns)
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Sheet {sheet_name} not found'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/delete_excel/<filename>', methods=['POST'])
def delete_excel(filename):
    filepath = os.path.join(EXCEL_FOLDER, filename)
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            flash(f"Excel file {filename} deleted successfully!", "success")
        except Exception as e:
            flash(f"Error deleting {filename}: {e}", "danger")
    else:
        flash(f"Excel file {filename} not found.", "danger")
    return redirect(request.referrer or url_for('output_excels'))

@app.route('/upload_template', methods=['POST'])
def upload_template():
    file = request.files.get('template_file')
    if file and file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        file.save(TEMPLATE_PATH)
        flash("Template uploaded successfully!", "success")
    else:
        flash("Please upload a valid Excel file.", "danger")
    return redirect(url_for('templates_page'))

@app.route('/delete_all_excels', methods=['POST'])
def delete_all_excels():
    errors = []
    deleted = 0
    if os.path.exists(EXCEL_FOLDER):
        for filename in os.listdir(EXCEL_FOLDER):
            filepath = os.path.join(EXCEL_FOLDER, filename)
            try:
                os.remove(filepath)
                deleted += 1
            except Exception as e:
                errors.append(f"Could not delete {filename}: {e}")
    if errors:
        return jsonify({'status': 'partial', 'errors': errors, 'deleted': deleted}), 207
    return jsonify({'status': 'success', 'deleted': deleted})

# @app.route('/upload', methods=['POST'])
# def upload_file():
#     file = request.files['file']
#     if file and file.filename:
#         filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
#         file.save(filepath)
#         flash(f"File {file.filename} uploaded successfully!", "success")
#     return redirect(url_for('extraction'))
from werkzeug.utils import secure_filename

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files.get('file')
    if file and file.filename:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        print("[DEBUG] Will save PDF to:", filepath)
        file.save(filepath)
        print("[DEBUG] Successfully saved PDF!")
        print("[DEBUG] Files in uploads:", os.listdir(app.config['UPLOAD_FOLDER']))
        flash(f"File {filename} uploaded successfully!", "success")
    else:
        print("[DEBUG] No file uploaded or missing filename")
        flash("No file selected!", "danger")
    return redirect(url_for('extraction'))

# @app.route('/process/<method>/<filename>', methods=['POST'])
# def process_file(method, filename):
#     if method == "flan":
#         from approaches.flan import extract_with_flan
#         result = extract_with_flan(filename)
#         if result.get('status') == 'success':
#             flash(f"Successfully processed {filename} using FLAN", "success")
#         else:
#             flash(f"Failed to process {filename} with FLAN", "danger")
#         return redirect(url_for('extraction'))
#     result = process_files([filename], method)
#     if result[filename]['status'] == 'success':
#         flash(f"Successfully processed {filename} using {method}", "success")
#     else:
#         flash(f"Failed to process {filename} with {method}", "danger")
#     return redirect(url_for('extraction'))
# @app.route('/process/<method>/<filename>', methods=['POST'])
# def process_file(method, filename):
#     ensure_filled_template()
#     if method == "flan":
#         from approaches.flan import extract_with_flan
#         result = extract_with_flan(filename)
#         # output_file = os.path.join(EXCEL_FOLDER, f'{os.path.splitext(filename)[0]}_extracted.xlsx')
#         # fill_template_all_sheets(output_file, FILLED_TEMPLATE_PATH)
#         output_file = os.path.join(EXCEL_FOLDER, f'{os.path.splitext(filename)[0]}_extracted.xlsx')
#         if not os.path.exists(output_file):
#             flash(f"Extraction failed: Output Excel was not created for {filename}.", "danger")
#             return redirect(url_for('extraction'))
#         fill_template_all_sheets(output_file, FILLED_TEMPLATE_PATH)
#         if result.get('status') == 'success':
#             flash(f"Successfully processed {filename} using FLAN", "success")
#             return send_file(FILLED_TEMPLATE_PATH, as_attachment=True)
#         else:
#             flash(f"Failed to process {filename} with FLAN", "danger")
#             return redirect(url_for('extraction'))
#     else:
#         result = process_files([filename], method)
#         if result[filename]['status'] == 'success':
#             flash(f"Successfully processed {filename} using {method}", "success")
#             output_files = [os.path.join(EXCEL_FOLDER, f) for f in os.listdir(EXCEL_FOLDER)
#                             if f.startswith(os.path.splitext(filename)[0])]
#             if output_files:
#                 output_file = max(output_files, key=os.path.getmtime)
#                 fill_template_all_sheets(output_file, FILLED_TEMPLATE_PATH)
#                 return send_file(FILLED_TEMPLATE_PATH, as_attachment=True)
#             else:
#                 flash("No output Excel file found after processing!", "danger")
#                 return redirect(url_for('extraction'))
#         else:
#             flash(f"Failed to process {filename} with {method}", "danger")
#             return redirect(url_for('extraction'))
@app.route('/process/<method>/<filename>', methods=['POST'])
def process_file(method, filename):
    ensure_filled_template()
    if method == "flan":
        from approaches.flan import extract_with_flan
        result = extract_with_flan(filename)
        output_file = os.path.join(EXCEL_FOLDER, f'{os.path.splitext(filename)[0]}_extracted.xlsx')
        # --- File existence check --- #
        if not os.path.exists(output_file):
            flash(f"Could not process {filename}: Excel extraction failed!", "danger")
            return redirect(url_for('extraction'))
        fill_template_all_sheets(output_file, FILLED_TEMPLATE_PATH)
        if result.get('status') == 'success':
            flash(f"Successfully processed {filename} using FLAN", "success")
            return send_file(FILLED_TEMPLATE_PATH, as_attachment=True)
        else:
            flash(f"Failed to process {filename} with FLAN", "danger")
            return redirect(url_for('extraction'))
    else:
        result = process_files([filename], method)
        if result[filename]['status'] == 'success':
            flash(f"Successfully processed {filename} using {method}", "success")
            output_files = [os.path.join(EXCEL_FOLDER, f) for f in os.listdir(EXCEL_FOLDER)
                            if f.startswith(os.path.splitext(filename)[0])]
            if output_files:
                output_file = max(output_files, key=os.path.getmtime)
                if not os.path.exists(output_file):
                    flash(f"Extraction failed: Output Excel was not created for {filename}.", "danger")
                    return redirect(url_for('extraction'))
                fill_template_all_sheets(output_file, FILLED_TEMPLATE_PATH)
                return send_file(FILLED_TEMPLATE_PATH, as_attachment=True)
            else:
                flash("No output Excel file found after processing!", "danger")
                return redirect(url_for('extraction'))
        else:
            flash(f"Failed to process {filename} with {method}", "danger")
            return redirect(url_for('extraction'))






@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/delete_upload/<filename>', methods=['POST'])
def delete_upload(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        flash(f"File {filename} deleted successfully!", "success")
    else:
        flash(f"File {filename} not found.", "danger")
    return redirect(request.referrer or url_for('directory'))

@app.route('/chat_with_bot', methods=['POST'])
def chat_with_bot():
    user_msg = request.get_json().get("message", "").strip()
    if not user_msg:
        return jsonify({"response": "Please enter a question."})
    answers = answer_question_per_pdf(user_msg, UPLOAD_FOLDER)
    if not answers:
        return jsonify({"response": "No PDF files found in the uploads folder."})
    combined = "\n".join(answers)
    return jsonify({"response": combined})

# Download filled template
@app.route('/download_filled_template')
def download_filled_template():
    if not os.path.exists(FILLED_TEMPLATE_PATH):
        flash("No filled template available to download.", "warning")
        return redirect(url_for('templates'))
    return send_file(FILLED_TEMPLATE_PATH, as_attachment=True)

# Delete filled template
@app.route('/delete_filled_template', methods=['POST'])
def delete_filled_template():
    if os.path.exists(FILLED_TEMPLATE_PATH):
        os.remove(FILLED_TEMPLATE_PATH)
        flash("Filled template deleted successfully.", "success")
    else:
        flash("No filled template found.", "warning")
    return redirect(url_for('templates'))

# Preview filled template (as HTML)
# def generate_filled_template_preview():
#     if os.path.exists(FILLED_TEMPLATE_PATH):
#         try:
#             df = pd.read_excel(FILLED_TEMPLATE_PATH)
#             return df.head(10).to_html(classes='table table-striped table-hover')
#         except Exception as e:
#             return f"Error loading preview: {e}"
#     return ""
def generate_filled_template_all_sheets_preview():
    import pandas as pd
    import os
    if not os.path.exists(FILLED_TEMPLATE_PATH):
        return ""
    try:
        xl = pd.ExcelFile(FILLED_TEMPLATE_PATH)
        tables_html = []
        for sheet in xl.sheet_names:
            df = xl.parse(sheet)
            # If you have huge sheets, change head() below to show only top N rows per sheet:
            html = df.to_html(classes="table table-bordered table-hover table-striped excel-preview-table", index=False, border=0)
            tables_html.append(f"<h5 style='margin-top:1.7em; color:#0291a7;'>{sheet}</h5>{html}")
        return "\n".join(tables_html)
    except Exception as e:
        return f"<div class='text-danger'>Could not load preview: {e}</div>"



# @app.route('/templates')
# def templates_page():
#     current_template = os.path.basename(TEMPLATE_PATH) if os.path.exists(TEMPLATE_PATH) else None
#     filled_template_exists = os.path.exists(FILLED_TEMPLATE_PATH)
#     filled_template_preview = generate_filled_template_preview() if filled_template_exists else ""
#     return render_template(
#         'templates.html',
#         current_template=current_template,
#         filled_template_exists=filled_template_exists,
#         filled_template_preview=filled_template_preview
#     )
@app.route('/templates')
def templates_page():
    current_template = os.path.basename(TEMPLATE_PATH) if os.path.exists(TEMPLATE_PATH) else None
    filled_template_exists = os.path.exists(FILLED_TEMPLATE_PATH)
    filled_template_preview = generate_filled_template_all_sheets_preview() if filled_template_exists else ""
    return render_template(
        'templates.html',
        current_template=current_template,
        filled_template_exists=filled_template_exists,
        filled_template_preview=filled_template_preview
    )


from shutil import copyfile

def ensure_filled_template():
    if not os.path.exists(FILLED_TEMPLATE_PATH):
        copyfile(TEMPLATE_PATH, FILLED_TEMPLATE_PATH)


if __name__ == '__main__':
    app.run(debug=False)
